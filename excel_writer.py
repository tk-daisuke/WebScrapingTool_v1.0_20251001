"""
Excel出力モジュール
スクレイピングしたデータをExcelファイルとして出力する
"""

import pandas as pd
import os
from datetime import datetime
import logging

# ログ設定
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelWriter:
    """Excelファイル出力を行うクラス"""
    
    def __init__(self, config):
        """
        初期化
        
        Args:
            config: configparserオブジェクト
        """
        self.config = config
        self.output_directory = config.get('Excel', 'output_directory', fallback='./output')
        
        # 出力ディレクトリが存在しない場合は作成
        os.makedirs(self.output_directory, exist_ok=True)
    
    def generate_filename(self):
        """
        タイムスタンプ付きのファイル名を生成する
        
        Returns:
            str: 生成されたファイル名
        """
        try:
            # 設定からファイル名テンプレートを取得
            filename_template = self.config.get('Excel', 'output_filename', 
                                               fallback='KPI_data_{timestamp}.xlsx')
            
            # タイムスタンプを生成
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # ファイル名を生成
            filename = filename_template.format(timestamp=timestamp)
            
            return filename
            
        except Exception as e:
            logger.error(f"ファイル名生成エラー: {e}")
            return f"KPI_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    def write_to_excel(self, dataframe, filename=None, sheet_name='KPIデータ'):
        """
        DataFrameをExcelファイルに出力する
        
        Args:
            dataframe (pd.DataFrame): 出力するデータ
            filename (str, optional): ファイル名（指定しない場合は自動生成）
            sheet_name (str): シート名
            
        Returns:
            str: 出力されたファイルのフルパス
        """
        try:
            if dataframe.empty:
                logger.warning("出力するデータが空です")
                return None
            
            # ファイル名を決定
            if filename is None:
                filename = self.generate_filename()
            
            # フルパスを生成
            filepath = os.path.join(self.output_directory, filename)
            
            # Excelファイルに出力
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # ワークシートを取得してフォーマット調整
                worksheet = writer.sheets[sheet_name]
                
                # 列幅を自動調整
                self._adjust_column_width(worksheet, dataframe)
                
                # ヘッダー行のスタイル設定
                self._format_header(worksheet)
            
            logger.info(f"Excelファイルを出力しました: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Excel出力エラー: {e}")
            return None
    
    def write_multiple_sheets(self, dataframes_dict, filename=None):
        """
        複数のDataFrameを複数シートのExcelファイルに出力する
        
        Args:
            dataframes_dict (dict): {シート名: DataFrame}の辞書
            filename (str, optional): ファイル名（指定しない場合は自動生成）
            
        Returns:
            str: 出力されたファイルのフルパス
        """
        try:
            if not dataframes_dict:
                logger.warning("出力するデータが空です")
                return None
            
            # ファイル名を決定
            if filename is None:
                filename = self.generate_filename()
            
            # フルパスを生成
            filepath = os.path.join(self.output_directory, filename)
            
            # Excelファイルに出力
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for sheet_name, df in dataframes_dict.items():
                    if not df.empty:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # ワークシートを取得してフォーマット調整
                        worksheet = writer.sheets[sheet_name]
                        self._adjust_column_width(worksheet, df)
                        self._format_header(worksheet)
            
            logger.info(f"複数シートExcelファイルを出力しました: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"複数シートExcel出力エラー: {e}")
            return None
    
    def _adjust_column_width(self, worksheet, dataframe):
        """
        列幅を自動調整する
        
        Args:
            worksheet: openpyxlワークシート
            dataframe: pandas DataFrame
        """
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # 列内の最大文字数を計算
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 列幅を設定（最小10、最大50）
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"列幅調整エラー: {e}")
    
    def _format_header(self, worksheet):
        """
        ヘッダー行のフォーマットを設定する
        
        Args:
            worksheet: openpyxlワークシート
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # ヘッダー行のスタイル
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 1行目（ヘッダー）にスタイルを適用
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
        except Exception as e:
            logger.warning(f"ヘッダーフォーマットエラー: {e}")
    
    def add_summary_sheet(self, filepath, summary_data):
        """
        既存のExcelファイルにサマリーシートを追加する
        
        Args:
            filepath (str): 既存のExcelファイルパス
            summary_data (dict): サマリーデータ
            
        Returns:
            bool: 成功の可否
        """
        try:
            from openpyxl import load_workbook
            
            # 既存のワークブックを読み込み
            workbook = load_workbook(filepath)
            
            # サマリーシートを作成
            summary_sheet = workbook.create_sheet("サマリー", 0)  # 最初のシートとして挿入
            
            # サマリーデータを書き込み
            row = 1
            for key, value in summary_data.items():
                summary_sheet.cell(row=row, column=1, value=key)
                summary_sheet.cell(row=row, column=2, value=value)
                row += 1
            
            # ファイルを保存
            workbook.save(filepath)
            
            logger.info("サマリーシートを追加しました")
            return True
            
        except Exception as e:
            logger.error(f"サマリーシート追加エラー: {e}")
            return False


def test_excel_writer():
    """Excel出力モジュールのテスト関数"""
    import configparser
    
    # テスト用設定
    config = configparser.ConfigParser()
    config.read('config.ini')
    
    # テストデータを作成
    test_data = {
        '日付': ['2024-01-01', '2024-01-02', '2024-01-03'],
        '売上高': [1000000, 1200000, 950000],
        '訪問者数': [5000, 6200, 4800],
        'コンバージョン率': [2.5, 3.1, 2.8]
    }
    
    df = pd.DataFrame(test_data)
    
    # Excel出力テスト
    writer = ExcelWriter(config)
    
    try:
        filepath = writer.write_to_excel(df, 'test_kpi_data.xlsx')
        
        if filepath and os.path.exists(filepath):
            print("Excel出力テスト成功")
            print(f"出力ファイル: {filepath}")
            
            # サマリーシート追加テスト
            summary_data = {
                '総売上': df['売上高'].sum(),
                '平均訪問者数': df['訪問者数'].mean(),
                '最高コンバージョン率': df['コンバージョン率'].max()
            }
            
            writer.add_summary_sheet(filepath, summary_data)
            print("サマリーシート追加完了")
            
        else:
            print("Excel出力テスト失敗")
            
    except Exception as e:
        print(f"テストエラー: {e}")


if __name__ == "__main__":
    test_excel_writer()
